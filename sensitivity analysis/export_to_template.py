import argparse
import json
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook


def load_config(config_path: Path) -> dict:
    with config_path.open("r", encoding="utf-8") as f:
        config = json.load(f)

    required_keys = [
        "t0",
        "start_date",
        "end_date",
        "shock_sheet",
        "template_after_sheet",
        "shock_file",
        "shock_types_file",
    ]
    missing = [k for k in required_keys if k not in config]
    if missing:
        raise KeyError(f"Missing keys in config: {missing}")

    return config


def read_sheet_as_records(path: Path, sheet_name: str) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        if not wb.sheetnames:
            raise ValueError(f"No sheets found in workbook: {path}")
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError(f"Sheet '{sheet_name}' in {path} is empty.") from exc

    headers = [str(v).strip() if v is not None else "" for v in header_row]
    if not any(headers):
        raise ValueError(f"Header row is empty in sheet '{sheet_name}' ({path}).")

    records: list[dict] = []
    for row in rows:
        if row is None:
            continue
        if all(v is None for v in row):
            continue
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else None
        records.append(record)

    wb.close()
    return headers, records


def get_period_columns(headers: list[str], t0: str, end_date: str) -> list[str]:
    if t0 not in headers:
        raise KeyError(f"t0 column '{t0}' not found in source data headers.")
    if end_date not in headers:
        raise KeyError(f"end_date column '{end_date}' not found in source data headers.")
    start_idx = headers.index(t0)
    end_idx = headers.index(end_date)
    if end_idx < start_idx:
        raise ValueError(f"end_date '{end_date}' appears before t0 '{t0}'.")
    return headers[start_idx : end_idx + 1]


def normalize(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def build_source_index(records: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for rec in records:
        mnemonic = normalize(rec.get("Mnemonic"))
        if not mnemonic:
            continue
        out.setdefault(mnemonic, []).append(rec)
    return out


def first_empty_column(ws) -> int:
    col = 1
    while ws.cell(row=1, column=col).value is not None:
        col += 1
    return col


def header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        name = normalize(ws.cell(row=1, column=col).value)
        if name:
            mapping[name] = col
    return mapping


def ordered_headers(ws) -> list[str]:
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        headers.append(normalize(ws.cell(row=1, column=col).value))
    return headers


def ensure_period_headers(ws, period_cols: list[str]) -> dict[str, int]:
    hm = header_map(ws)
    for name in period_cols:
        if name in hm:
            continue
        col = first_empty_column(ws)
        ws.cell(row=1, column=col, value=name)
        hm[name] = col
    return hm


def select_record(candidates: list[dict], shock_change_tested: str) -> Optional[dict]:
    if not candidates:
        return None
    if len(candidates) == 1 or not shock_change_tested:
        return candidates[0]
    for rec in candidates:
        if normalize(rec.get("Shock Change Tested")) == shock_change_tested:
            return rec
    return candidates[0]


def write_row(ws, row_idx: int, column_map: dict[str, int], record: dict, write_columns: list[str]) -> None:
    for col_name in write_columns:
        if col_name not in column_map:
            continue
        ws.cell(row=row_idx, column=column_map[col_name], value=record.get(col_name))


def fill_template_sheet(
    ws,
    source_records: list[dict],
    period_cols: list[str],
) -> tuple[int, int]:
    src_by_mnemonic = build_source_index(source_records)
    hm = ensure_period_headers(ws, period_cols)

    if "Mnemonic" not in hm:
        raise KeyError(f"Sheet '{ws.title}' is missing required header 'Mnemonic'.")

    template_headers = [normalize(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    write_columns = [h for h in template_headers if h and h in source_records[0]] if source_records else []

    # Ensure period columns are always included.
    for col_name in period_cols:
        if col_name not in write_columns:
            write_columns.append(col_name)

    wrote = 0
    missed = 0
    existing_rows = ws.max_row
    has_template_data_rows = existing_rows > 1

    if has_template_data_rows:
        for row_idx in range(2, existing_rows + 1):
            mnemonic = normalize(ws.cell(row=row_idx, column=hm["Mnemonic"]).value)
            if not mnemonic:
                continue
            shock_val = normalize(ws.cell(row=row_idx, column=hm.get("Shock Change Tested", 0)).value) if "Shock Change Tested" in hm else ""
            candidates = src_by_mnemonic.get(mnemonic, [])
            rec = select_record(candidates, shock_val)
            if rec is None:
                missed += 1
                continue
            write_row(ws, row_idx, hm, rec, write_columns)
            wrote += 1
    else:
        row_idx = 2
        for rec in source_records:
            write_row(ws, row_idx, hm, rec, write_columns)
            row_idx += 1
            wrote += 1

    return wrote, missed


def to_number(value, sheet_name: str, row_idx: int, col_name: str) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Non-numeric value in sheet '{sheet_name}', row {row_idx}, column '{col_name}': {value!r}"
        ) from exc


def load_pct_adjustment_rules(
    shock_types_path: Path,
) -> tuple[dict[str, float], list[dict]]:
    with shock_types_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    adjustments = payload.get("adjustments")
    if not isinstance(adjustments, dict):
        raise KeyError("Missing or invalid 'adjustments' in shock types file.")

    decline_map = adjustments.get("decline_pct_of_t0")
    if not isinstance(decline_map, dict):
        raise KeyError("Missing or invalid 'adjustments.decline_pct_of_t0' in shock types file.")
    decline_pct_by_shock: dict[str, float] = {}
    for k, v in decline_map.items():
        if not normalize(k):
            continue
        if not isinstance(v, (int, float)):
            raise TypeError(
                f"'adjustments.decline_pct_of_t0.{k}' must be numeric."
            )
        decline_pct_by_shock[normalize(k)] = float(v)

    fx_rules = adjustments.get("pct_of_t0_add_by_mnemonic", [])
    if not isinstance(fx_rules, list):
        raise KeyError(
            "Missing or invalid 'adjustments.pct_of_t0_add_by_mnemonic' in shock types file."
        )
    normalized_fx_rules: list[dict] = []
    for idx, rule in enumerate(fx_rules):
        if not isinstance(rule, dict):
            raise TypeError(
                f"'adjustments.pct_of_t0_add_by_mnemonic[{idx}]' must be an object."
            )
        shock_label = normalize(rule.get("shock_label"))
        mnemonics = rule.get("mnemonics")
        pct_of_t0_add = rule.get("pct_of_t0_add")
        if not shock_label:
            raise ValueError(
                f"'adjustments.pct_of_t0_add_by_mnemonic[{idx}].shock_label' is required."
            )
        if not isinstance(mnemonics, list) or not all(
            isinstance(m, str) for m in mnemonics
        ):
            raise TypeError(
                f"'adjustments.pct_of_t0_add_by_mnemonic[{idx}].mnemonics' must be a list of strings."
            )
        if not isinstance(pct_of_t0_add, (int, float)):
            raise TypeError(
                f"'adjustments.pct_of_t0_add_by_mnemonic[{idx}].pct_of_t0_add' must be numeric."
            )
        normalized_fx_rules.append(
            {
                "shock_label": shock_label,
                "mnemonics": {normalize(m) for m in mnemonics if normalize(m)},
                "pct_of_t0_add": float(pct_of_t0_add),
            }
        )
    return decline_pct_by_shock, normalized_fx_rules


def get_metric_mode(
    shock_label: str,
    mnemonic: str,
    decline_pct_by_shock: dict[str, float],
    fx_rules: list[dict],
) -> Optional[str]:
    for rule in fx_rules:
        if rule["shock_label"] == shock_label and mnemonic in rule["mnemonics"]:
            return "min" if float(rule["pct_of_t0_add"]) < 0 else "max"
    if shock_label in decline_pct_by_shock:
        return "min"
    return None


def generate_comparison_workbook(
    before_ws,
    after_ws,
    t0: str,
    start_date: str,
    end_date: str,
    decline_pct_by_shock: dict[str, float],
    fx_rules: list[dict],
    comparison_path: Path,
) -> tuple[int, int]:
    before_headers = ordered_headers(before_ws)
    after_headers = ordered_headers(after_ws)
    period_cols = get_period_columns(after_headers, t0, end_date)
    start_to_end_cols = get_period_columns(after_headers, start_date, end_date)

    before_hm = header_map(before_ws)
    after_hm = header_map(after_ws)
    for col_name in period_cols:
        if col_name not in before_hm:
            raise KeyError(
                f"Column '{col_name}' not found in before sheet '{before_ws.title}'."
            )
        if col_name not in after_hm:
            raise KeyError(
                f"Column '{col_name}' not found in after sheet '{after_ws.title}'."
            )
    for col_name in start_to_end_cols:
        if col_name not in before_hm:
            raise KeyError(
                f"Column '{col_name}' not found in before sheet '{before_ws.title}'."
            )
        if col_name not in after_hm:
            raise KeyError(
                f"Column '{col_name}' not found in after sheet '{after_ws.title}'."
            )
    if "Mnemonic" not in before_hm or "Mnemonic" not in after_hm:
        raise KeyError("Both comparison sheets must contain 'Mnemonic'.")

    has_shock = "Shock Change Tested" in before_hm and "Shock Change Tested" in after_hm

    before_by_full: dict[tuple[str, str], int] = {}
    before_by_mnemonic: dict[str, int] = {}
    for row_idx in range(2, before_ws.max_row + 1):
        mnemonic = normalize(before_ws.cell(row=row_idx, column=before_hm["Mnemonic"]).value)
        if not mnemonic:
            continue
        shock = (
            normalize(before_ws.cell(row=row_idx, column=before_hm["Shock Change Tested"]).value)
            if has_shock
            else ""
        )
        before_by_full.setdefault((mnemonic, shock), row_idx)
        before_by_mnemonic.setdefault(mnemonic, row_idx)

    comp_wb = Workbook()
    comp_ws = comp_wb.active
    comp_ws.title = "Comparison"

    # Keep the same column layout as the after sheet.
    for idx, col_name in enumerate(after_headers, start=1):
        if col_name:
            comp_ws.cell(row=1, column=idx, value=col_name)
    extra_start_col = len(after_headers) + 1
    comp_ws.cell(row=1, column=extra_start_col, value="before_shock")
    comp_ws.cell(row=1, column=extra_start_col + 1, value="after_shock")
    comp_ws.cell(row=1, column=extra_start_col + 2, value="diff")

    written = 0
    unmatched = 0
    out_row = 2
    for after_row_idx in range(2, after_ws.max_row + 1):
        mnemonic = normalize(after_ws.cell(row=after_row_idx, column=after_hm["Mnemonic"]).value)
        if not mnemonic:
            continue
        shock = (
            normalize(after_ws.cell(row=after_row_idx, column=after_hm["Shock Change Tested"]).value)
            if has_shock
            else ""
        )
        before_row_idx = before_by_full.get((mnemonic, shock))
        if before_row_idx is None:
            before_row_idx = before_by_mnemonic.get(mnemonic)
        if before_row_idx is None:
            unmatched += 1
            continue

        for col_idx, col_name in enumerate(after_headers, start=1):
            if not col_name:
                continue
            if col_name in period_cols:
                after_val = after_ws.cell(row=after_row_idx, column=after_hm[col_name]).value
                before_val = before_ws.cell(
                    row=before_row_idx, column=before_hm[col_name]
                ).value
                diff = to_number(
                    after_val, after_ws.title, after_row_idx, col_name
                ) - to_number(before_val, before_ws.title, before_row_idx, col_name)
                comp_ws.cell(row=out_row, column=col_idx, value=diff)
            else:
                comp_ws.cell(
                    row=out_row,
                    column=col_idx,
                    value=after_ws.cell(row=after_row_idx, column=after_hm[col_name]).value,
                )

        if has_shock:
            metric_mode = get_metric_mode(
                shock, mnemonic, decline_pct_by_shock, fx_rules
            )
        else:
            metric_mode = None
        if metric_mode is not None:
            before_t0 = to_number(
                before_ws.cell(row=before_row_idx, column=before_hm[t0]).value,
                before_ws.title,
                before_row_idx,
                t0,
            )
            after_t0 = to_number(
                after_ws.cell(row=after_row_idx, column=after_hm[t0]).value,
                after_ws.title,
                after_row_idx,
                t0,
            )
            if before_t0 == 0.0 or after_t0 == 0.0:
                raise ValueError(
                    f"Cannot compute additional decline metrics with t0=0 for mnemonic '{mnemonic}'."
                )

            before_window = [
                to_number(
                    before_ws.cell(row=before_row_idx, column=before_hm[col_name]).value,
                    before_ws.title,
                    before_row_idx,
                    col_name,
                )
                for col_name in start_to_end_cols
            ]
            after_window = [
                to_number(
                    after_ws.cell(row=after_row_idx, column=after_hm[col_name]).value,
                    after_ws.title,
                    after_row_idx,
                    col_name,
                )
                for col_name in start_to_end_cols
            ]
            if metric_mode == "min":
                before_extreme = min(before_window)
                after_extreme = min(after_window)
            else:
                before_extreme = max(before_window)
                after_extreme = max(after_window)
            before_shock = before_extreme / before_t0 - 1.0
            after_shock = after_extreme / after_t0 - 1.0
            shock_diff = after_shock - before_shock
            comp_ws.cell(row=out_row, column=extra_start_col, value=before_shock)
            comp_ws.cell(row=out_row, column=extra_start_col + 1, value=after_shock)
            comp_ws.cell(row=out_row, column=extra_start_col + 2, value=shock_diff)

        out_row += 1
        written += 1

    comparison_path.parent.mkdir(parents=True, exist_ok=True)
    comp_wb.save(comparison_path)
    comp_wb.close()
    return written, unmatched


def run(
    config_path: Path,
    template_path: Optional[Path],
    before_path: Path,
    after_path: Path,
    out_path: Optional[Path],
    comparison_path: Optional[Path],
) -> None:
    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")
    if not before_path.exists():
        raise FileNotFoundError(f"Missing source file: {before_path}")
    if not after_path.exists():
        raise FileNotFoundError(f"Missing source file: {after_path}")

    config = load_config(config_path)
    config_base = config_path.resolve().parent
    shock_types_path = config_base / config["shock_types_file"]
    if not shock_types_path.exists():
        raise FileNotFoundError(f"Shock types file not found: {shock_types_path}")
    decline_pct_by_shock, fx_rules = load_pct_adjustment_rules(shock_types_path)
    if template_path is None:
        template_file = config.get("template_file", config["shock_file"])
        template_path = config_base / template_file
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    if out_path is None:
        shock_stem = Path(config["shock_file"]).stem
        out_path = config_base.parent / "output" / f"{shock_stem}_distribution.xlsx"
    if comparison_path is None:
        comparison_path = out_path.parent / "comparison.xlsx"
    before_headers, before_records = read_sheet_as_records(before_path, "Sheet1")
    after_headers, after_records = read_sheet_as_records(after_path, "Sheet1")

    period_cols_before = get_period_columns(before_headers, config["t0"], config["end_date"])
    period_cols_after = get_period_columns(after_headers, config["t0"], config["end_date"])

    wb = load_workbook(template_path)
    before_sheet = config["shock_sheet"]
    after_sheet = config["template_after_sheet"]
    if before_sheet not in wb.sheetnames:
        raise KeyError(f"Template is missing sheet '{before_sheet}'.")
    if after_sheet not in wb.sheetnames:
        raise KeyError(f"Template is missing sheet '{after_sheet}'.")

    wrote_before, missed_before = fill_template_sheet(
        wb[before_sheet], before_records, period_cols_before
    )
    wrote_after, missed_after = fill_template_sheet(
        wb[after_sheet], after_records, period_cols_after
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    comp_rows, comp_unmatched = generate_comparison_workbook(
        wb[before_sheet],
        wb[after_sheet],
        config["t0"],
        config["start_date"],
        config["end_date"],
        decline_pct_by_shock,
        fx_rules,
        comparison_path,
    )
    wb.close()

    print(f"Saved template output: {out_path}")
    print(f"Saved comparison output: {comparison_path}")
    print(f"{before_sheet}: wrote {wrote_before} rows; missed {missed_before} template mnemonics.")
    print(f"{after_sheet}: wrote {wrote_after} rows; missed {missed_after} template mnemonics.")
    print(
        f"Comparison: wrote {comp_rows} rows; unmatched rows (after vs before): {comp_unmatched}."
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Write Extra_shock_before/after data into input.xlsx template tabs."
    )
    parser.add_argument(
        "--config",
        default="input/sensitivity_config.json",
        help="Path to config JSON (uses t0/end_date).",
    )
    parser.add_argument(
        "--template",
        default=None,
        help="Path to template workbook. If omitted, uses shock_file in config.",
    )
    parser.add_argument(
        "--before",
        default="output/Extra_shock_before.xlsx",
        help="Path to before-adjustment workbook.",
    )
    parser.add_argument(
        "--after",
        default="output/Extra_shock_after.xlsx",
        help="Path to after-adjustment workbook.",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Path for filled template workbook. If omitted, uses output/<shock_file>_distribution.xlsx.",
    )
    parser.add_argument(
        "--comparison-out",
        default=None,
        help="Path for comparison workbook. If omitted, uses output/comparison.xlsx.",
    )
    args = parser.parse_args()

    run(
        Path(args.config),
        Path(args.template) if args.template else None,
        Path(args.before),
        Path(args.after),
        Path(args.out) if args.out else None,
        Path(args.comparison_out) if args.comparison_out else None,
    )


if __name__ == "__main__":
    main()
