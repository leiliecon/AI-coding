import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"

TEMPLATE_PATH = INPUT_DIR / "template_input.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "path2shock_finalV.xlsx"
EXPORT_RULES_PATH = INPUT_DIR / "export_rules.json"

DEFAULT_SCENARIO_ALIASES = {}
DEFAULT_M_FILL_MODES = {}
DEFAULT_FILL_MODE = "single_shock"
VALID_FILL_MODES = {
    "single_shock",
    "two_row_shock_then_extreme",
    "two_row_extreme_then_shock",
    "range_extreme_to_shock",
}


def _norm_text(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def _build_output_suffix_map(output_dir: Path):
    suffix_to_path = {}
    for file_path in output_dir.glob("path2shock_*.xlsx"):
        if file_path.name.lower() == OUTPUT_FILE.name.lower():
            continue
        stem = file_path.stem
        suffix = stem.replace("path2shock_", "", 1)
        suffix_to_path[_norm_text(suffix)] = file_path
    return suffix_to_path


def _load_export_rules(rules_path: Path = EXPORT_RULES_PATH):
    scenario_aliases = dict(DEFAULT_SCENARIO_ALIASES)
    m_fill_modes = dict(DEFAULT_M_FILL_MODES)
    default_fill_mode = DEFAULT_FILL_MODE

    if not rules_path.exists():
        raise FileNotFoundError(f"Export rules file not found: {rules_path}")

    with open(rules_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    user_aliases = raw.get("scenario_aliases", {})
    user_m_fill_modes = raw.get("m_fill_modes", {})
    user_default_fill_mode = raw.get("default_fill_mode")

    if isinstance(user_aliases, dict):
        scenario_aliases.update(
            {_norm_text(k): _norm_text(v) for k, v in user_aliases.items()}
        )
    if isinstance(user_m_fill_modes, dict):
        normalized_modes = {
            str(k).strip().upper(): str(v).strip()
            for k, v in user_m_fill_modes.items()
        }
        m_fill_modes.update(normalized_modes)
    if user_default_fill_mode is not None:
        default_fill_mode = str(user_default_fill_mode).strip()

    for m_name, fill_mode in m_fill_modes.items():
        if fill_mode not in VALID_FILL_MODES:
            raise ValueError(
                f"Invalid fill mode '{fill_mode}' for M name '{m_name}'. "
                f"Valid modes: {sorted(VALID_FILL_MODES)}"
            )
    if default_fill_mode not in VALID_FILL_MODES:
        raise ValueError(
            f"Invalid default fill mode '{default_fill_mode}'. "
            f"Valid modes: {sorted(VALID_FILL_MODES)}"
        )

    return scenario_aliases, m_fill_modes, default_fill_mode


def _resolve_scenario_path(scenario_name: str, suffix_to_path, scenario_aliases):
    scenario_key = _norm_text(scenario_name)
    if not scenario_key:
        return None

    candidates = {scenario_key}

    alias_key = scenario_aliases.get(scenario_key)
    if alias_key:
        candidates.add(_norm_text(alias_key))

    # Also support reverse aliasing so mappings work in either direction.
    for src, dst in scenario_aliases.items():
        src_n = _norm_text(src)
        dst_n = _norm_text(dst)
        if scenario_key == src_n:
            candidates.add(dst_n)
        if scenario_key == dst_n:
            candidates.add(src_n)

    for key in candidates:
        if key in suffix_to_path:
            return suffix_to_path[key]

    for suffix, path in suffix_to_path.items():
        for key in candidates:
            if key.startswith(suffix) or suffix.startswith(key):
                return path

    return None


def _load_scenario_data(path: Path):
    df = pd.read_excel(path)
    required_cols = {"M names", "Slides name", "shock", "extreme_level"}
    missing = required_cols - set(df.columns)
    if missing:
        missing_str = ", ".join(sorted(missing))
        raise ValueError(f"Missing required columns in {path.name}: {missing_str}")

    by_slide = {}
    by_m_name = {}
    for _, row in df.iterrows():
        slide_name = row["Slides name"]
        slide_key = _norm_text(slide_name)
        m_name = str(row["M names"]).strip() if pd.notna(row["M names"]) else ""
        m_key = _norm_text(m_name)
        payload = {
            "m_name": m_name,
            "shock": row["shock"],
            "extreme_level": row["extreme_level"],
        }
        if not slide_key:
            if m_key:
                by_m_name[m_key] = payload
            continue
        by_slide[slide_key] = payload
        if m_key:
            by_m_name[m_key] = payload
    return {"by_slide": by_slide, "by_m_name": by_m_name}


def export_path2shock_table(
    template_path: Path = TEMPLATE_PATH,
    output_dir: Path = OUTPUT_DIR,
    output_file: Path = OUTPUT_FILE,
):
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = wb.active
    scenario_aliases, m_fill_modes, default_fill_mode = _load_export_rules()

    suffix_to_path = _build_output_suffix_map(output_dir)
    if not suffix_to_path:
        raise FileNotFoundError(
            f"No output files found under {output_dir}. Expected path2shock_*.xlsx"
        )

    scenario_data_by_col = {}
    for col_idx in range(2, ws.max_column + 1):
        scenario_cell_value = ws.cell(row=2, column=col_idx).value
        if scenario_cell_value is None or str(scenario_cell_value).strip() == "":
            continue

        scenario_path = _resolve_scenario_path(
            str(scenario_cell_value), suffix_to_path, scenario_aliases
        )
        if scenario_path is None:
            raise ValueError(
                f"Could not find output file for scenario '{scenario_cell_value}'"
            )

        scenario_data_by_col[col_idx] = _load_scenario_data(scenario_path)

    # Template rows start from row 3 (row 1 display, row 2 scenario names).
    for row_idx in range(3, ws.max_row + 1):
        slide_value = ws.cell(row=row_idx, column=1).value
        if slide_value is None or str(slide_value).strip() == "":
            continue

        slide_key = _norm_text(slide_value)

        for col_idx, scenario_maps in scenario_data_by_col.items():
            by_slide = scenario_maps["by_slide"]
            by_m_name = scenario_maps["by_m_name"]
            row_data = by_slide.get(slide_key) or by_m_name.get(slide_key)
            if not row_data:
                continue

            m_name = row_data["m_name"].upper()
            shock_val = row_data["shock"]
            extreme_val = row_data["extreme_level"]
            fill_mode = m_fill_modes.get(m_name, default_fill_mode)

            if fill_mode == "two_row_extreme_then_shock":
                if pd.notna(extreme_val):
                    ws.cell(row=row_idx, column=col_idx, value=extreme_val)
                if row_idx + 1 <= ws.max_row and pd.notna(shock_val):
                    ws.cell(row=row_idx + 1, column=col_idx, value=shock_val)
            elif fill_mode == "two_row_shock_then_extreme":
                if pd.notna(shock_val):
                    ws.cell(row=row_idx, column=col_idx, value=shock_val)
                if row_idx + 1 <= ws.max_row and pd.notna(extreme_val):
                    ws.cell(row=row_idx + 1, column=col_idx, value=extreme_val)
            elif fill_mode == "range_extreme_to_shock":
                if (
                    pd.notna(extreme_val)
                    and pd.notna(shock_val)
                ):
                    ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=f"{extreme_val} to {shock_val}",
                    )
            elif fill_mode == "single_shock":
                if pd.notna(shock_val):
                    ws.cell(row=row_idx, column=col_idx, value=shock_val)

    center_alignment = Alignment(horizontal="center", vertical="center")
    final_font = Font(name="Inter", size=11, color="FF002060")
    white_font = Font(name="Inter", size=11, color="FFFFFFFF", bold=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                cell.alignment = center_alignment
                if cell.row <= 2:
                    cell.font = white_font
                else:
                    cell.font = final_font

    wb.save(output_file)
    return output_file


if __name__ == "__main__":
    saved_path = export_path2shock_table()
    print(f"Saved: {saved_path}")
